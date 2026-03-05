from pathlib import Path

from openpyxl import load_workbook


def parse(path: Path, max_pages: int | None = None) -> str:
    """Parse XLSX files and extract cell data from all sheets."""
    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []

    sheets = wb.sheetnames
    if max_pages is not None:
        sheets = sheets[:max_pages]

    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

    wb.close()

    if not parts:
        return "(No data could be extracted from this XLSX)"

    return "\n\n".join(parts)
